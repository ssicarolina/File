# pembagian kelompok
from openpyxl import Workbook
from openpyxl import load_workbook
from random import randint

wb = load_workbook(filename = "DPK-KU1102-02-sem1-thn2019.xlsx")
sh = wb["KU1102-02"]

T = []
n = 10
while True:
    nim = sh["B"+str(n)].value
    if nim == None:
        break
    T.append(nim)
    n = n+1

N_mhs = len(T)
sisa = N_mhs % 5
N_kelompok = (N_mhs-sisa)//5

Kelompok = {}
for i in range(0, N_kelompok):
    if sisa > 0:
        N = 6
        sisa = sisa-1
    else:
        N = 5
    Kelompok[i+1] = []
    for j in range(0, N):
        Kelompok[i+1].append(T.pop(randint(0, len(T)-1)))

wb2 = Workbook()
ws = wb2.active
ws.title = "Kelompok TB 1"
alphabet = "ABCDEFGHIJKLMNOPQRSTUVWXYZ"
for k in Kelompok.keys():
    ws[alphabet[k-1] + "1"] = "KEL " + str(k)
    for j in range(0, len(Kelompok[k])):
        ws[alphabet[k-1] + str(j+2)] = str(Kelompok[k][j])[-3:]
wb2.save("Kelompok.xlsx")

